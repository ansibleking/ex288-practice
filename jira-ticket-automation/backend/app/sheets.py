from __future__ import annotations

import csv
import io
from datetime import date, datetime, time

from openpyxl import load_workbook
from pydantic import BaseModel

# Hard caps so a huge/pathological upload can't blow up memory or produce a
# table the browser can't render -- this is a manual visualization aid, not a
# spreadsheet engine, so truncating extreme sheets is an acceptable tradeoff.
MAX_ROWS = 2000
MAX_COLS = 100


class SheetParseError(ValueError):
    """Raised when the uploaded file can't be parsed as a spreadsheet."""


class SheetData(BaseModel):
    name: str
    headers: list[str]
    rows: list[list[str]]
    truncated: bool


class ParsedWorkbook(BaseModel):
    filename: str
    sheets: list[SheetData]


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _parse_xlsx(content: bytes) -> list[SheetData]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises several distinct exception types for bad files
        raise SheetParseError(f"Could not read this as an Excel file: {exc}") from exc

    sheets: list[SheetData] = []
    for worksheet in workbook.worksheets:
        all_rows = list(worksheet.iter_rows(values_only=True))
        truncated = len(all_rows) > MAX_ROWS + 1 or (
            all_rows and len(all_rows[0]) > MAX_COLS
        )
        header_row = all_rows[0] if all_rows else ()
        headers = [_cell_to_str(c) or f"Column {i + 1}" for i, c in enumerate(header_row[:MAX_COLS])]
        rows = [
            [_cell_to_str(c) for c in row[:MAX_COLS]]
            for row in all_rows[1 : MAX_ROWS + 1]
        ]
        sheets.append(SheetData(name=worksheet.title, headers=headers, rows=rows, truncated=truncated))
    workbook.close()
    return sheets


def _parse_csv(content: bytes) -> list[SheetData]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    truncated = len(all_rows) > MAX_ROWS + 1 or (all_rows and len(all_rows[0]) > MAX_COLS)
    header_row = all_rows[0] if all_rows else []
    headers = [c or f"Column {i + 1}" for i, c in enumerate(header_row[:MAX_COLS])]
    rows = [row[:MAX_COLS] for row in all_rows[1 : MAX_ROWS + 1]]
    return [SheetData(name="Sheet1", headers=headers, rows=rows, truncated=truncated)]


def parse_spreadsheet(filename: str, content: bytes) -> ParsedWorkbook:
    lower = filename.lower()
    if lower.endswith(".csv"):
        sheets = _parse_csv(content)
    elif lower.endswith(".xlsx"):
        sheets = _parse_xlsx(content)
    else:
        raise SheetParseError("Only .xlsx and .csv files are supported.")
    return ParsedWorkbook(filename=filename, sheets=sheets)
