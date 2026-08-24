import io

import pytest
from openpyxl import Workbook

from app.sheets import MAX_COLS, MAX_ROWS, SheetParseError, parse_spreadsheet


def _xlsx_bytes(rows: list[list[object]], sheet_name: str = "Sheet1") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def test_parse_xlsx_reads_headers_and_rows():
    content = _xlsx_bytes(
        [
            ["Host", "IP", "Approved"],
            ["web01", "10.0.0.1", True],
            ["web02", "10.0.0.2", False],
        ]
    )

    result = parse_spreadsheet("network-access.xlsx", content)

    assert result.filename == "network-access.xlsx"
    assert len(result.sheets) == 1
    sheet = result.sheets[0]
    assert sheet.name == "Sheet1"
    assert sheet.headers == ["Host", "IP", "Approved"]
    assert sheet.rows == [["web01", "10.0.0.1", "True"], ["web02", "10.0.0.2", "False"]]
    assert sheet.truncated is False


def test_parse_xlsx_handles_multiple_sheets():
    workbook = Workbook()
    workbook.active.title = "Servers"
    workbook.active.append(["Host"])
    workbook.active.append(["web01"])
    workbook.create_sheet("Ports").append(["Port"])
    buf = io.BytesIO()
    workbook.save(buf)

    result = parse_spreadsheet("multi.xlsx", buf.getvalue())

    assert [s.name for s in result.sheets] == ["Servers", "Ports"]


def test_parse_xlsx_truncates_oversized_sheets():
    rows = [["Host"]] + [[f"host{i}"] for i in range(MAX_ROWS + 50)]
    content = _xlsx_bytes(rows)

    result = parse_spreadsheet("big.xlsx", content)

    sheet = result.sheets[0]
    assert sheet.truncated is True
    assert len(sheet.rows) == MAX_ROWS


def test_parse_xlsx_truncates_wide_sheets():
    header = [f"col{i}" for i in range(MAX_COLS + 10)]
    content = _xlsx_bytes([header, ["v"] * len(header)])

    result = parse_spreadsheet("wide.xlsx", content)

    sheet = result.sheets[0]
    assert sheet.truncated is True
    assert len(sheet.headers) == MAX_COLS


def test_parse_csv_reads_headers_and_rows():
    content = b"Host,IP\nweb01,10.0.0.1\nweb02,10.0.0.2\n"

    result = parse_spreadsheet("network-access.csv", content)

    assert len(result.sheets) == 1
    sheet = result.sheets[0]
    assert sheet.headers == ["Host", "IP"]
    assert sheet.rows == [["web01", "10.0.0.1"], ["web02", "10.0.0.2"]]


def test_parse_rejects_unsupported_extension():
    with pytest.raises(SheetParseError, match="Only .xlsx and .csv"):
        parse_spreadsheet("notes.txt", b"hello")


def test_parse_xlsx_rejects_garbage_bytes():
    with pytest.raises(SheetParseError, match="Could not read"):
        parse_spreadsheet("fake.xlsx", b"not a real xlsx file")
